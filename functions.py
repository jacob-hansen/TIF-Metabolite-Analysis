from libraries import *
from standards import *
from openpyxl import load_workbook

def requestDocs():
    """
    Requests input from the user to find the excel files.
    """
    count = 0
    found = False

    # Look for documents
    while(not found):
        cwd = os.getcwd()
        excelDocs = glob.glob(cwd+'/*.xlsx',recursive=True)
        excelDocs = glob.glob(cwd+'/*.xlsx',recursive=True)

        if(len(excelDocs) > 0):
            print(excelDocs[0])
            for i in excelDocs:
                print("One excel document found was "+str(i.rpartition('/')[-1]))
                answer = False
                answer = input("Is this the correct excel document (y/n)?")
                while(answer != 'y' and answer != 'n'):
                    answer = input("Did not understand input. Please enter y for yes or n for no:")
                if answer == 'y':
                    excelDocs = [str(i)]
                    print(i)
                    found = True
                    break
                else:
                    answer = False

        if(len(excelDocs) == 0 or answer == False):
            print('No .xlsx documents were found in the coding folder, please move you file there and try again')
            sys.exit()

    # if (len(excelDocs) == 1):
    #     break
        #Code below can be used to collect only 1 excel file:
        # elif(len(excelDocs) > 1):
        #     cfile = input("It appears there are multiple excel files in that folder. Please enter the file name: ")
        #     count = 0
        #     while(True):
        #         excelDocs = glob.glob(cwd+'/'+cfolder+'/**/'+cfile+'.XLS',recursive=True)
        #         if (len(excelDocs) == 1):
        #             break
        #         elif(len(excelDocs) > 1):
        #             print("It appears there are multiple excel files in that folder. Please enter the file name: ")
        #         else:
        #             print("It appears that the folder is empty or mistyped. Try again")
        #         if(count > 0):
        #             cfile = input('File not found, remember not to include the .XLS. Please try again or quit the program with "Control C": ')
        #         else:
        #             cfile = input("File not found, please try again: ")
        #             count += 1
        #    break
        # else:
        #     if(count > 0):
        #         print('The program still cannot find the specified folder. Please try again or quit with "Control C"')
        #     elif(len(excelDocs > 1)):
        #         print('The program found multiple files in the folder, please remove all other files or add the file name with folder/filename:')
        #     else:
        #         print("It appears that the folder is empty or mistyped. Try again")
        #         count += 1
    #excelDocs.sort()
    fileName = input("Please enter master file name: ")
    mPath = cwd+'/'+fileName+'.xlsx'
    return  (mPath, excelDocs, fileName)

def readDocs(excelDocs):
    """
    List of documents to check as Input
    Collects the information into
    """

    sheetName = 'response ratios'
    fileTest = pd.read_excel(excelDocs[0], sheet_name = sheetName)
    fileTest.index.name = None

    columns = list(fileTest.columns)
    fileFinal = fileTest.set_index(fileTest[columns[1]], drop = False)

    # identify incorrectly named compounds and attempt to find them in the standards
    toChange = {}
    for key in columns:
        if(key not in iStandards):
            word = key.replace(" ", "-")
            wordReformat = word.lower()[0].upper()+word.lower()[1:]
            if(word in iStandards):
                toChange[key] = word
            elif(wordReformat in iStandards):
                toChange[key] = wordReformat
            elif(word.lower() in iStandards):
                toChange[key] = word.lower()
            if(word.upper() in iStandards):
                toChange[key] = word.upper()
            elif(word[0].lower()+key[1:] in iStandards):
                toChange[key] = word[0].lower()+key[1:]
            elif(word[0].upper()+key[1:] in iStandards):
                toChange[key] = word[0].upper()+key[1:]

    # Pull data until the first empty row
    indices = fileFinal.index.tolist()
    indChange = {}
    for ind in indices:
        test = detectStandard(ind)
        if(test != "empty"):
            indChange[ind] = test
    fileFinal = fileFinal.rename(index=indChange)

    #Changing the names of compounds in case they were formatted differently
    for item in toChange:
        position = columns.index(item)
        columns[position] = toChange[item]
        fileFinal[toChange[item]] = fileFinal.pop(item)

    return (fileFinal)


def findCompoundStatus(mPath, fileName, values):
    with pd.ExcelWriter((fileName+'.xlsx'), engine='xlsxwriter') as writer:
        workbook = writer.book
        cSheetName = 'Internal-External Input'
        worksheet = workbook.add_worksheet(cSheetName)
        writer.sheets[cSheetName] = worksheet

        columns = list(values.columns)
        worksheet.write_column(1, 0, columns)
        listToWrite = []
        for item in columns:
            if (item in iStandards):
                listToWrite.append('i')
            else:
                listToWrite.append('s')
        worksheet.write_column(1, 1, listToWrite)

        sampleSort = ['N', 'Normal', 'T', 'P', 'Tumor']
        worksheet.write_column(1, 2, sampleSort)
        cName = 'Compound'
        c2Name = 'Internal or external (i, e, or s to skip compound)'
        worksheet.write_row(0, 0, [cName, c2Name, 'Sorting Keys'])


        column = 0
        worksheet = workbook.add_worksheet('DFS Response')
        writer.sheets['DFS Response'] = worksheet

        values.to_excel(writer, sheet_name='DFS Response',
                startrow=0, startcol=column, index = True)

    time.sleep(2)
    os.system('open '+fileName+'.xlsx')
    print("Please edit files by changing all external standards to 'e' and then saving and closing the file. ")
    while(True):
        test = input("Please enter 'done' when finished: ")
        if (test == 'done'):
            print("Thank you!")
            break
        else:
            print("Command not understood. ")
    shutil.move((fileName+'.xlsx'), mPath)
    dfcStatus = pd.read_excel(mPath, usecols = [cName, c2Name], header = 0, sheet_name = cSheetName)
    sortKeys = pd.read_excel(mPath, usecols = ['Sorting Keys'], header = 0, sheet_name = cSheetName).dropna()
    sortKeys = sortKeys['Sorting Keys'].to_list()
    compoundStatus = dfcStatus.to_dict()
    compoundStatus = dict((compoundStatus[cName].get(k, k), v) for (k, v) in compoundStatus[c2Name].items())
    return compoundStatus, sortKeys

def detectStandard(name):
    """
    given a name
    returns empty if the name is not a standard,
    otherwise returns the name of the standard
    """
    posKeys = {"100nM": "0.1", "0.1": "0.1", "300": "0.3", "0.3": "0.3", "1uM": "1", "3uM": "3", "10uM": "10", "30uM": "30", "100uM": "100", "500": "500"}
    if("pool" not in name and "Pool" not in name):
        return("empty")
    for set in posKeys:
        if(set in name):
            return posKeys[set] #This will detect everytime it is done with uM or nM labeling and everything but 1, 10, 100, 3, and 30!!
    if("100" in name):
        return "100"
    elif("10" in name):
        return "10"
    elif("0nM" in name):
        return "0"
    elif("0uM" in name):
        return "0"
    elif("30" in name):
        return "30"
    else:
        numbers = re.findall('[0-9]+', str)
        if(numbers[-1] == 1):
            return "1"
        else:
            return "3"
    return("empty")

def removeSpecialChar(key):
    """
    Some names come with a special character that needs to be removed.
    """
    pos = 0
    keyErrorList = []
    storageString = "empty"
    for c in key:
        if (ord(c) < 0 or ord(c) > 127):
            keyErrorList.append(pos)
        pos += 1
    if(len(keyErrorList) != 0):
        storageStr = [char for char in key]
        storageStr1 = ''.join(storageStr[:min(keyErrorList)-1])
        storageStr2 = ''.join(storageStr[max(keyErrorList)+1:])
        storageString = ''.join([storageStr1, storageStr2])
    return(storageString)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    """
    Given a filename, dataframe, and where to add the dataframe in the excel sheet.
    Appends dataframe to the bottom of the excel sheet.
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def append_list_to_excel(filename, list, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    writer.write_column(0, 0, masterIndex)
    #df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()




def checkIsIn(check, recievedList):
    """
    Checking if check is in the recievedList
    Then finding the maximum length object with the name in it.
    """
    check = str(check[::-1])
    isInList = []
    notIn = False

    for i in range(len(recievedList)-1):
        isIn = recievedList[i][::-1]
        if(len(check) < len(isIn)):
            notIn = True
        else:
            for i in range(len(isIn)):
                if(check[i]!=isIn[i]):
                    notIn = True
        if(notIn == False):
            isInList.append(i)
        notIn = False
    if(len(isInList) == 0):
        return 'zzzzzzz'
    returnValue = isInList[0]

    for i in isInList:
        try:
            if(len(recievedList[i]) > len(recievedList[returnValue])):
                returnValue = i
        except(IndexError):
            pass

    return returnValue
